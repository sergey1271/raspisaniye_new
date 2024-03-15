import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.datavalidation import DataValidation
from database_create import get_classes, get_subjects, get_days_object, get_times, get_classrooms, get_teachers, add_studyplan, get_groups_ids_by_name, get_subject_ids_by_name, get_teachers_ids_by_name
DATABASE = "1234.db"


def create_wb(name: str) -> None:
    wb = openpyxl.Workbook()
    wb.save(f'data/{name}')

def create_sheet(name: str, sheet: str):
    wb = openpyxl.load_workbook(filename = f'data/{name}')
    wb.create_sheet(sheet)
    try:
        del wb["Sheet"]
    except:
        pass
    wb.save(f'data/{name}')

def write_classes(name: str, sheet_to_write: str, classes: list) -> None:
    wb = openpyxl.load_workbook(filename = f'data/{name}')
    sheet = wb[sheet_to_write]
    x0 = 3
    y0 = 3
    for cl in classes:
        sheet.cell(row=y0, column=x0).value = cl.class_name
        sheet.cell(row=y0, column=x0).border = Border(right=Side(border_style='medium', color='FF000000'))
        sheet.cell(row=y0, column=x0+1).border = Border(right=Side(border_style='medium', color='FF000000'))
        y0 += 1
    wb.save(f'data/{name}')

def write_subjects(name: str, sheet_to_write: str, subjects: list):
    wb = openpyxl.load_workbook(filename = f'data/{name}')
    sheet = wb[sheet_to_write]
    x0 = 5
    y0 = 2
    for sb in subjects:
        sheet.cell(row=y0, column=x0).value = sb.subject_name
        sheet.cell(row=y0, column=x0).alignment = Alignment(textRotation=90)
        sheet.cell(row=y0, column=x0).border = Border(bottom=Side(border_style='medium', color='FF000000'))
        x0 += 1
    wb.save(f'data/{name}')

def create_file_to_user(name: str, database: str):
    create_wb(name)
    uch_plan = "Учебный план"
    uchitelya = "Учителя"
    create_sheet(name, uch_plan)
    create_sheet(name, uchitelya)
    classes = get_classes(database)
    subjects = get_subjects(database)
    write_classes(name, uch_plan, classes)
    write_subjects(name, uch_plan, subjects)
    write_classes(name, uchitelya, classes)
    write_subjects(name, uchitelya, subjects)
    wb = openpyxl.load_workbook(filename = f'data/{name}')
    sheet = wb[uch_plan]
    y0 = 2
    x0 = 4
    sheet.cell(row=y0, column=x0).value = "Кол-во уроков в день макс."
    sheet.cell(row=y0, column=x0).alignment = Alignment(textRotation=90)
    sheet.cell(row=y0, column=x0).border = Border(bottom=Side(border_style='medium', color='FF000000'))
    sheet = wb[uchitelya]
    sheet.column_dimensions['A'].hidden= True  # скрываю столбец со списком учителей
    y00 = 2
    for tch in get_teachers(database, True):  # записываю список учителей
        sheet.cell(row=y00, column=1).value = tch.teacher_name
        y00 += 1
    data_val = DataValidation(type="list",formula1='=$A:$A', showErrorMessage=True) #You can change =$A:$A with a smaller range like =A1:A9
    sheet.add_data_validation(data_val)
    for i in range(len(classes)):
        for j in range(len(subjects)):
            data_val.add(sheet.cell(row=y0+1+i, column=x0+1+j))  # добавлюя проверку данных
    sheet.cell(row=y0, column=x0).value = "Классный руководитель"
    sheet.cell(row=y0, column=x0).alignment = Alignment(textRotation=90)
    sheet.cell(row=y0, column=x0).border = Border(bottom=Side(border_style='medium', color='FF000000'))
    wb.save(f'data/{name}')

def read_file(name: str, database: str):
    uch_plan = "Учебный план"
    uchitelya = "Учителя"
    groups = get_classes(database)
    subjects = get_subjects(database)
    wb = openpyxl.load_workbook(filename = f'{name}')
    groups_ids = get_groups_ids_by_name(database)
    subjects_ids = get_subject_ids_by_name(database)
    teachers_ids = get_teachers_ids_by_name(database)
    y0, x0 = 3, 5
    for gr in range(y0, y0+len(groups)):
        for sb in range(x0, x0+len(subjects)):
            sheet = wb[uch_plan]
            lessons = sheet.cell(gr, sb).value
            if lessons == 0 or lessons == None:
                continue
            print(sheet.cell(gr, x0-2).value)
            group = groups_ids[sheet.cell(gr, x0-2).value]
            subject = subjects_ids[sheet.cell(y0-1, sb).value]
            sheet = wb[uchitelya]
            teacher = teachers_ids[sheet.cell(gr, sb).value]
            add_studyplan(database, group, lessons, teacher, subject)
    

def format_sheet(name: str, sheet_to_write: str, database: str, lessons: int):
    x0 = 3
    y0 = 9
    x00 = x0
    y00 = y0
    wb = openpyxl.load_workbook(filename = f'data/{name}')
    sheet = wb[sheet_to_write]
    # Оглавление таблицы
    sheet.merge_cells(start_row=y0-4, start_column=x0, end_row=y0-3, end_column=x0)
    sheet.cell(row=y0-4, column=x0).value = "День\nнедели" 
    sheet.cell(row=y0-4, column=x0).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    
    sheet.merge_cells(start_row=y0-4, start_column=x0+1, end_row=y0-3, end_column=x0+1)
    sheet.cell(row=y0-4, column=x0+1).value = "Урок"   
    sheet.cell(row=y0-4, column=x0+1).alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.merge_cells(start_row=y0-4, start_column=x0+2, end_row=y0-3, end_column=x0+2)
    sheet.cell(row=y0-4, column=x0+2).value = "Время"   
    sheet.cell(row=y0-4, column=x0+2).alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.cell(row=y0-2, column=x0+2).value = "Класс"   
    sheet.cell(row=y0-2, column=x0+2).alignment = Alignment(horizontal='center', vertical='center')
    
    sheet.cell(row=y0-1, column=x0+2).value = "Подгруппа"   
    sheet.cell(row=y0-1, column=x0+2).alignment = Alignment(horizontal='center', vertical='center')
    
    times_to_excel = dict()
    for d in get_days_object(database):
        sheet.merge_cells(start_row=y0, start_column=x0, end_row=y0+2*lessons-1, end_column=x0)
        sheet.cell(row=y0, column=x0).value = d.day_name
        sheet.cell(row=y0, column=x0).alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        sheet.cell(row=y0, column=x0).font = Font(size=18, bold=True)
        y01 = y0
        for t in get_times(database)[d.day_id]:
            sheet.merge_cells(start_row=y01, start_column=x0+1, end_row=y01+1, end_column=x0+1)
            sheet.merge_cells(start_row=y01, start_column=x0+2, end_row=y01+1, end_column=x0+2)
            sheet.cell(row=y01, column=x0+1).value = f"{t.period}"
            sheet.cell(row=y01, column=x0+2).value = f"{t.start}\n - {t.end}"
            sheet.cell(row=y01, column=x0+2).font = Font(bold=True)
            sheet.cell(row=y01, column=x0+1).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=y01, column=x0+2).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            sheet.cell(row=y01, column=x0+2).border = Border(right=Side(border_style='thin', color='FF000000'))
            
            times_to_excel[t.time_id] = y01 
            y01 += 2
        y0 += 2*lessons
        
    x0 = x00
    y0 = y00
    classes_to_excel = dict()
    for cl in get_classes(database):
        sheet.merge_cells(start_row=y0-4, start_column=x0+3, end_row=y0-4, end_column=x0+5)
        sheet.cell(row=y0-4, column=x0+3).value = "Предмет"
        sheet.cell(row=y0-4, column=x0+6).value = "Ауд."
        sheet.cell(row=y0-4, column=x0+6).alignment = Alignment(vertical="center", horizontal="right")
        sheet.merge_cells(start_row=y0-3, start_column=x0+3, end_row=y0-3, end_column=x0+6)
        sheet.cell(row=y0-3, column=x0+3).value = "Учитель"
        sheet.merge_cells(start_row=y0-2, start_column=x0+3, end_row=y0-2, end_column=x0+6)
        classes_to_excel[cl.class_id] = x0+3
        sheet.cell(row=y0-2, column=x0+3).value = cl.class_name
        sheet.cell(row=y0-2, column=x0+3).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=y0-2, column=x0+3).font = Font(bold=True)
        
        sheet.merge_cells(start_row=y0-1, start_column=x0+3, end_row=y0-1, end_column=x0+4)
        sheet.merge_cells(start_row=y0-1, start_column=x0+5, end_row=y0-1, end_column=x0+6)
        sheet.cell(row=y0-1, column=x0+3).border = Border(bottom=Side(border_style='thick', color='FF000000'))
        sheet.cell(row=y0-1, column=x0+5).border = Border(bottom=Side(border_style='thick', color='FF000000'))
        
        y_t = y0-1
        for d in get_days_object(database):
            y_tt = y_t + 2
            for x in range(lessons-1):
                sheet.cell(row=y_tt, column=x0+3).border = Border(bottom=Side(border_style='thin', color='FF000000'))
                sheet.cell(row=y_tt, column=x0+4).border = Border(bottom=Side(border_style='thin', color='FF000000'))
                sheet.cell(row=y_tt, column=x0+5).border = Border(bottom=Side(border_style='thin', color='FF000000'))
                sheet.cell(row=y_tt, column=x0+6).border = Border(right=Side(border_style='thin', color='FF000000'), \
                                                                    bottom=Side(border_style='thin', color='FF000000'))
                sheet.cell(row=y_tt-1, column=x0+6).border = Border(right=Side(border_style='thin', color='FF000000'))
                y_tt += 2
            sheet.cell(row=y_tt, column=x0+6).border = Border(right=Side(border_style='thin', color='FF000000'))
            sheet.cell(row=y_tt-1, column=x0+6).border = Border(right=Side(border_style='thin', color='FF000000'))
            y_t += 2*lessons
            sheet.cell(row=y_t, column=x0+3).border = Border(bottom=Side(border_style='thick', color='FF000000'))
            sheet.cell(row=y_t, column=x0+4).border = Border(bottom=Side(border_style='thick', color='FF000000'))
            sheet.cell(row=y_t, column=x0+5).border = Border(bottom=Side(border_style='thick', color='FF000000'))
            sheet.cell(row=y_t, column=x0+6).border = Border(bottom=Side(border_style='thick', color='FF000000'), \
                                                                right=Side(border_style='thin', color='FF000000'))
        x0 += 4
    wb.save(f'data/{name}')
    
    return classes_to_excel, times_to_excel

def export_timetable(entity: list, name: str, database: str):
    e0 = entity[0]  # lesson / time
    e1 = entity[1]  # lesson / classroom
    # cоздам файл и листы
    create_wb(name)
    create_sheet(name, "Расписание")
    classes_to_excel, times_to_excel = format_sheet(name, "Расписание", database, 7)  # Добавить чтение кол-ва уроков
    
    wb = openpyxl.load_workbook(filename = f'data/{name}')
    sheet = wb["Расписание"]
    for l in e0.keys():
        for gr in l.groups:
            x0 = classes_to_excel[gr]
            for t in range(len(e0[l])):
                y0 = times_to_excel[e0[l][t].id]
                subject = next(subject for subject in get_subjects(database) if subject.subject_id == l.subject)
                sheet.merge_cells(start_row=y0, start_column=x0, end_row=y0, end_column=x0+2)
                sheet.cell(row=y0, column=x0).value = subject.subject_name
                classroom = next(clrm for clrm in get_classrooms(database) if clrm.classroom_id == e1[l][t].id)
                sheet.cell(row=y0, column=x0+3).value = classroom.number
                sheet.cell(row=y0, column=x0+3).alignment = Alignment(vertical="center", horizontal="right")
                sheet.merge_cells(start_row=y0+1, start_column=x0, end_row=y0+1, end_column=x0+3)
                sheet.cell(row=y0+1, column=x0).value = get_teachers(database)[l.teacher.id].teacher_name
               
    wb.save(f'data/{name}')

# create_file_to_user("1234.xlsx", DATABASE)
# read_file(f'1234.xlsx', 'test.db')
#export_timetable([1, 2, 3], "12345.xlsx", DATABASE)