from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

ws = wb.create_sheet('New Sheet')

for number in range(1,100): #Generates 99 "ip" address in the Column A;
    ws['A{}'.format(number)].value= "192.168.1.{}".format(number)

data_val = DataValidation(type="list",formula1='=$A:$A') #You can change =$A:$A with a smaller range like =A1:A9
ws.add_data_validation(data_val)

data_val.add(ws["B1"]) #If you go to the cell B1 you will find a drop down list with all the values from the column A

wb.save('Test.xlsx')